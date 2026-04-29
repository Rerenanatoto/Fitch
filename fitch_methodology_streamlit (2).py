# ============================================================
# Fitch Sovereign Methodology + Global Sovereign Data Comparator
# Streamlit App – v4 (Dashboard e Tabela inspirados no S&P)
# ============================================================

import io
import math
import re
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="Fitch Sovereign Methodology + Comparator", layout="wide")

INTERCEPT = 4.877

# ============================================================
# Metadata – SRM  (inalterado do v3)
# ============================================================

SRM_VARIABLES = {
    "structural": {
        "governance_indicator": {
            "label": "Composite governance indicator (percentile rank)",
            "coefficient": 0.079, "weight": 22.1,
            "help": "Simple average percentile rank of World Bank governance indicators.",
        },
        "gdp_per_capita_percentile": {
            "label": "GDP per capita percentile rank (USD, market FX)",
            "coefficient": 0.037, "weight": 11.7,
            "help": "Percentile rank across Fitch-rated sovereigns.",
        },
        "share_world_gdp_log": {
            "label": "Share in world GDP (natural log of % share)",
            "coefficient": 0.645, "weight": 14.5,
            "help": "Natural logarithm of the percentage share in world GDP.",
        },
        "years_since_default_transform": {
            "label": "Years since default/restructuring event (SRM transformed value)",
            "coefficient": -1.744, "weight": 4.3,
            "help": "Use 0 if no event after 1980, or input the SRM-ready transformed value.",
        },
        "money_supply_log": {
            "label": "Broad money supply (% of GDP) - natural log",
            "coefficient": 0.148, "weight": 1.1,
            "help": "Natural logarithm of broad money as % of GDP.",
        },
    },
    "macro": {
        "real_gdp_growth_volatility_log": {
            "label": "Real GDP growth volatility - natural log",
            "coefficient": -0.704, "weight": 4.5,
            "help": "Natural log of the exponentially weighted std. deviation of historical real GDP growth.",
        },
        "consumer_price_inflation": {
            "label": "Consumer price inflation (3-year centred average, %, truncated 2%-50%)",
            "coefficient": -0.068, "weight": 3.6,
            "help": "Public criteria truncate this variable to the 2%-50% range.",
        },
        "real_gdp_growth": {
            "label": "Real GDP growth (3-year centred average, %)",
            "coefficient": 0.054, "weight": 1.7,
            "help": "Three-year centred average.",
        },
    },
    "public_finances": {
        "gross_general_govt_debt": {
            "label": "Gross general government debt (% of GDP, 3-year centred average)",
            "coefficient": -0.023, "weight": 9.2,
            "help": "Gross general government debt, centred three-year average.",
        },
        "general_govt_interest_revenue": {
            "label": "General government interest (% of revenues, 3-year centred average)",
            "coefficient": -0.044, "weight": 4.6,
            "help": "General government interest expenditures as % of revenues.",
        },
        "general_govt_fiscal_balance": {
            "label": "General government fiscal balance (% of GDP, 3-year centred average)",
            "coefficient": 0.039, "weight": 2.1,
            "help": "General government fiscal balance, centred three-year average.",
        },
        "fc_govt_debt_share": {
            "label": "Foreign-currency government debt (% of gross government debt, 3-yr centred avg)",
            "coefficient": -0.008, "weight": 3.2,
            "help": "Foreign-currency or indexed government debt as % of gross government debt.",
        },
    },
    "external": {
        "reserve_currency_flexibility": {
            "label": "Reserve-currency flexibility (SRM transformed value)",
            "coefficient": 0.484, "weight": 7.1,
            "help": "Use 0 if the sovereign has no reserve-currency flexibility.",
        },
        "sovereign_net_foreign_assets": {
            "label": "Sovereign net foreign assets (% of GDP, 3-year centred average)",
            "coefficient": 0.010, "weight": 7.5,
            "help": "Three-year centred average of sovereign net foreign assets.",
        },
        "commodity_dependence": {
            "label": "Commodity dependence (% of current external receipts)",
            "coefficient": -0.003, "weight": 1.0,
            "help": "Non-manufactured merchandise exports as % of current external receipts.",
        },
        "fx_reserves_months_cxp": {
            "label": "Official FX reserves (months of CXP) - for non-reserve-currency sovereigns",
            "coefficient": 0.021, "weight": 1.2,
            "help": "Usually set to 0 if reserve-currency flexibility is above 0.",
        },
        "external_interest_service": {
            "label": "External interest service (% of current external receipts, 3-yr centred avg)",
            "coefficient": -0.004, "weight": 0.2,
            "help": "Three-year centred average.",
        },
        "cab_plus_net_fdi": {
            "label": "Current account balance + net inward FDI (% of GDP, 3-yr centred avg)",
            "coefficient": 0.004, "weight": 0.4,
            "help": "Three-year centred average.",
        },
    },
}

PILLAR_LABELS = {
    "structural": "I. Structural Features",
    "macro": "II. Macroeconomic Performance, Policies and Prospects",
    "public_finances": "III. Public Finances",
    "external": "IV. External Finances",
}

QO_FACTORS = {
    "structural": [
        "Political stability and capacity",
        "Financial sector risks",
        "Other structural factors",
    ],
    "macro": [
        "Macroeconomic policy credibility and flexibility",
        "GDP growth outlook",
        "Macroeconomic stability / imbalances",
    ],
    "public_finances": [
        "Fiscal financing flexibility",
        "Public debt sustainability",
        "Fiscal structure",
    ],
    "external": [
        "External financing flexibility",
        "External debt sustainability",
        "Vulnerability to shocks",
    ],
}

QO_GUIDANCE = {
    2: "Exceptionally strong features relative to SRM data and output",
    1: "Strong features relative to SRM data and output",
    0: "Average features relative to SRM data and output",
    -1: "Weak features relative to SRM data and output",
    -2: "Exceptionally weak features relative to SRM data and output",
}

RATING_SCALE_NUMERIC = {
    16: "AAA", 15: "AA+", 14: "AA", 13: "AA-",
    12: "A+", 11: "A", 10: "A-",
    9: "BBB+", 8: "BBB", 7: "BBB-",
    6: "BB+", 5: "BB", 4: "BB-",
    3: "B+", 2: "B", 1: "B-",
}

LONG_TERM_SCALE = [
    "AAA", "AA+", "AA", "AA-", "A+", "A", "A-",
    "BBB+", "BBB", "BBB-", "BB+", "BB", "BB-",
    "B+", "B", "B-", "CCC+", "CCC", "CCC-", "CC", "C", "RD", "D",
]

ST_MAPPING_OPTIONS = {
    "AAA": ("F1+", "F1+"), "AA+": ("F1+", "F1+"), "AA": ("F1+", "F1+"), "AA-": ("F1+", "F1+"),
    "A+": ("F1", "F1+"), "A": ("F1", "F1+"), "A-": ("F2", "F1"),
    "BBB+": ("F2", "F1"), "BBB": ("F3", "F2"), "BBB-": ("F3", "F3"),
    "BB+": ("B", "B"), "BB": ("B", "B"), "BB-": ("B", "B"),
    "B+": ("B", "B"), "B": ("B", "B"), "B-": ("B", "B"),
    "CCC+": ("C", "C"), "CCC": ("C", "C"), "CCC-": ("C", "C"),
    "CC": ("C", "C"), "C": ("C", "C"),
    "RD": ("C/RD/D", "C/RD/D"), "D": ("D", "D"),
}

VARIABLE_RULES = {
    "governance_indicator": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "gdp_per_capita_percentile": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "share_world_gdp_log": {"step": 0.1, "soft_min": -20, "soft_max": 5},
    "years_since_default_transform": {"step": 0.01, "soft_min": 0, "soft_max": 5},
    "money_supply_log": {"step": 0.1, "soft_min": -5, "soft_max": 10},
    "real_gdp_growth_volatility_log": {"step": 0.1, "soft_min": -5, "soft_max": 10},
    "consumer_price_inflation": {"step": 0.1, "soft_min": -100, "soft_max": 500},
    "real_gdp_growth": {"step": 0.1, "soft_min": -20, "soft_max": 20},
    "gross_general_govt_debt": {"step": 0.1, "soft_min": 0, "soft_max": 500},
    "general_govt_interest_revenue": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "general_govt_fiscal_balance": {"step": 0.1, "soft_min": -50, "soft_max": 50},
    "fc_govt_debt_share": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "reserve_currency_flexibility": {"step": 0.1, "soft_min": 0, "soft_max": 20},
    "sovereign_net_foreign_assets": {"step": 0.1, "soft_min": -300, "soft_max": 300},
    "commodity_dependence": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "fx_reserves_months_cxp": {"step": 0.1, "soft_min": 0, "soft_max": 60},
    "external_interest_service": {"step": 0.1, "soft_min": 0, "soft_max": 100},
    "cab_plus_net_fdi": {"step": 0.1, "soft_min": -30, "soft_max": 30},
}

# ============================================================
# Helpers gerais
# ============================================================

def normalize_label(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r"\s+", " ", text)
    return text

def slugify(text: str) -> str:
    text = normalize_label(text).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text

def coerce_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.replace({
        "N/A": np.nan, "N.M.": np.nan, "NM": np.nan,
        "None": np.nan, "nan": np.nan, "": np.nan,
        "..": np.nan, "...": np.nan, "n.a": np.nan,
        "n.a.": np.nan, "-": np.nan,
    })
    return pd.to_numeric(s, errors="coerce")

def st_dataframe_compat(df: pd.DataFrame, **kwargs):
    try:
        st.dataframe(df, **kwargs)
    except TypeError:
        st.dataframe(df)

def st_plotly_chart_compat(fig, use_container_width: bool = True):
    try:
        st.plotly_chart(fig, use_container_width=use_container_width)
    except TypeError:
        st.plotly_chart(fig)

# ============================================================
# SRM Helpers  (inalterado do v3)
# ============================================================

def clamp_qo(adjustments: dict, crisis_extension: bool = False) -> int:
    raw = int(sum(adjustments.values()))
    return raw if crisis_extension else max(-3, min(3, raw))

def score_to_lt_rating(score: float) -> str:
    rounded = int(round(score))
    if rounded >= 16:
        return "AAA"
    if rounded <= 0:
        return "CCC+"
    return RATING_SCALE_NUMERIC.get(rounded, "CCC+")

def rating_index(rating: str) -> int:
    return LONG_TERM_SCALE.index(rating) if rating in LONG_TERM_SCALE else LONG_TERM_SCALE.index("CCC+")

def apply_notches(base_rating: str, notch_adjustment: int) -> str:
    if base_rating not in LONG_TERM_SCALE:
        return base_rating
    idx = rating_index(base_rating)
    new_idx = max(0, min(len(LONG_TERM_SCALE) - 1, idx - int(notch_adjustment)))
    return LONG_TERM_SCALE[new_idx]

def map_short_term(long_rating: str, use_higher_option: bool) -> str:
    lo, hi = ST_MAPPING_OPTIONS.get(long_rating, ("C", "C"))
    return hi if use_higher_option else lo

def approx_years_since_default_transform(years_since_event=None, no_event_since_1980=True) -> float:
    if no_event_since_1980 or years_since_event is None:
        return 0.0
    years_since_event = max(0.0, float(years_since_event))
    return float(math.exp(-math.log(2) * years_since_event / 4.3))

def safe_number_input(var_key: str, label: str, default: float, help_text: str):
    rule = VARIABLE_RULES.get(var_key, {"step": 0.1})
    return st.number_input(
        label, value=float(default), step=float(rule.get("step", 0.1)),
        key=var_key, help=help_text,
    )

def get_clean_srm_inputs() -> dict:
    inputs = {}
    for pillar in SRM_VARIABLES.values():
        for k in pillar.keys():
            v = float(st.session_state.get(k, 0.0))
            if k == "consumer_price_inflation":
                v = min(50.0, max(2.0, v))
            inputs[k] = v
    return inputs

def compute_srm(inputs: dict) -> tuple:
    details = []
    total = INTERCEPT
    for pillar_key, vars_dict in SRM_VARIABLES.items():
        for var_key, meta in vars_dict.items():
            value = float(inputs.get(var_key, 0.0))
            contribution = value * float(meta["coefficient"])
            total += contribution
            details.append({
                "Pillar": PILLAR_LABELS[pillar_key],
                "Variable": meta["label"],
                "Value": value,
                "Coefficient": meta["coefficient"],
                "Contribution": contribution,
            })
    details.append({
        "Pillar": "Intercept", "Variable": "OLS intercept",
        "Value": 1.0, "Coefficient": INTERCEPT, "Contribution": INTERCEPT,
    })
    return total, details

def build_radar(srm_score: float, qo_total: int, final_score: float):
    categories = ["SRM score", "QO total", "Final model score"]
    values = [srm_score, qo_total, final_score]
    categories += categories[:1]
    values += values[:1]
    fig = go.Figure(data=[go.Scatterpolar(r=values, theta=categories, fill="toself")])
    fig.update_layout(showlegend=False, polar=dict(radialaxis=dict(visible=True)))
    return fig

# ============================================================
# State init  (inalterado do v3)
# ============================================================

def init_state():
    defaults = {
        "governance_indicator": 50.0,
        "gdp_per_capita_percentile": 50.0,
        "share_world_gdp_log": -4.0,
        "years_since_default_transform": 0.0,
        "money_supply_log": math.log(60.0),
        "real_gdp_growth_volatility_log": math.log(3.0),
        "consumer_price_inflation": 4.0,
        "real_gdp_growth": 3.0,
        "gross_general_govt_debt": 60.0,
        "general_govt_interest_revenue": 8.0,
        "general_govt_fiscal_balance": -3.0,
        "fc_govt_debt_share": 35.0,
        "reserve_currency_flexibility": 0.0,
        "sovereign_net_foreign_assets": -20.0,
        "commodity_dependence": 25.0,
        "fx_reserves_months_cxp": 4.0,
        "external_interest_service": 8.0,
        "cab_plus_net_fdi": -1.0,
        "qo_structural": 0,
        "qo_macro": 0,
        "qo_public_finances": 0,
        "qo_external": 0,
        "qo_crisis_extension": False,
        "lc_manual_adjust": 0,
        "fc_robust_liquidity": False,
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)

# ============================================================
# Fitch XLSB Parser  (inalterado do v3)
# ============================================================

@st.cache_data(show_spinner=False)
def parse_fitch_comparator(file_bytes) -> pd.DataFrame:
    """Parse the Fitch Global Sovereign Data Comparator .xlsb format."""
    from pyxlsb import open_workbook

    wb = open_workbook(io.BytesIO(file_bytes))
    all_rows = []
    sheet_name = wb.sheets[0] if wb.sheets else None
    if sheet_name is None:
        return pd.DataFrame()

    with wb.get_sheet(sheet_name) as sheet:
        for row in sheet.rows():
            all_rows.append([c.v for c in row])

    if not all_rows:
        return pd.DataFrame()

    raw = pd.DataFrame(all_rows)

    # 1. Find the "period row" – the one with many year-like values (2020, 2021, …)
    cat_row_idx = None
    for i in range(min(30, len(raw))):
        row_str = " ".join([str(v) for v in raw.iloc[i].tolist() if v is not None])
        if re.search(r"\b20(2[0-9]|3[0-9])\b", row_str) and row_str.count("20") > 5:
            cat_row_idx = i
            break

    if cat_row_idx is None:
        st.error("Não foi possível localizar a linha de períodos no XLSB.")
        return pd.DataFrame()

    # 2. Extract header rows
    section_row_idx = max(0, cat_row_idx - 4)
    indicator_row_idx = max(0, cat_row_idx - 2)
    unit_row_idx = max(0, cat_row_idx - 1)
    period_row_idx = cat_row_idx

    ncols = len(raw.columns)

    def get_row_vals(idx):
        vals = []
        for c in range(ncols):
            v = raw.iloc[idx, c] if idx < len(raw) else None
            vals.append(str(v).strip() if v is not None else "")
        return vals

    sections_raw = get_row_vals(section_row_idx)
    indicators_raw = get_row_vals(indicator_row_idx)
    units_raw = get_row_vals(unit_row_idx)
    periods_raw = get_row_vals(period_row_idx)

    def forward_fill(lst):
        result, current = [], ""
        for v in lst:
            if v and v.lower() not in ("none", "nan", ""):
                current = v
            result.append(current)
        return result

    sections = forward_fill(sections_raw)
    indicators = forward_fill(indicators_raw)

    # 3. Data rows start after period row
    data_start = cat_row_idx + 1

    # 4. Find first data column
    meta_end = 0
    for c in range(ncols):
        p = periods_raw[c].strip()
        if re.match(r"^(19|20)\d{2}$", p) or "av." in p.lower() or "latest" in p.lower():
            meta_end = c
            break
    if meta_end == 0:
        meta_end = 14

    # 5. Build column metadata
    col_meta = []
    for c in range(meta_end, ncols):
        section = normalize_label(sections[c]) if c < len(sections) else ""
        indicator = normalize_label(indicators[c]) if c < len(indicators) else ""
        unit = normalize_label(units_raw[c]) if c < len(units_raw) else ""
        period = normalize_label(periods_raw[c]) if c < len(periods_raw) else ""

        if not section and not indicator:
            continue
        if not period or period.lower() in ("none", "nan", ""):
            continue

        year_match = re.search(r"(19|20)\d{2}", period)
        year_num = int(year_match.group()) if year_match else None
        is_average = "av." in period.lower() or "average" in period.lower()
        is_forecast = year_num is not None and year_num >= 2025
        is_latest = "latest" in period.lower()

        col_meta.append({
            "col_idx": c,
            "section": section,
            "indicator": indicator,
            "unit": unit,
            "year": period,
            "year_num": year_num if year_num else 0,
            "is_average": is_average,
            "is_forecast": is_forecast,
            "is_latest": is_latest,
        })

    if not col_meta:
        st.error("Não foi possível mapear colunas de dados.")
        return pd.DataFrame()

    # 6. Parse data rows
    records = []
    for r in range(data_start, len(raw)):
        row = raw.iloc[r]
        entity_key = str(row.iloc[0]).strip() if row.iloc[0] is not None else ""
        if not entity_key or entity_key.lower() in ("none", "nan", ""):
            continue

        entity_type_raw = str(row.iloc[4]).strip().upper() if len(row) > 4 and row.iloc[4] is not None else ""
        if "COUNTRY" not in entity_type_raw and "HEADING" not in entity_type_raw:
            if not any(c.isalpha() for c in entity_key):
                continue

        country_name = normalize_label(str(row.iloc[1])) if len(row) > 1 and row.iloc[1] is not None else entity_key
        country_code = normalize_label(str(row.iloc[2])) if len(row) > 2 and row.iloc[2] is not None else ""
        lt_fc_rating = normalize_label(str(row.iloc[3])) if len(row) > 3 and row.iloc[3] is not None else ""
        entity_type = "COUNTRY" if "COUNTRY" in entity_type_raw else "GROUP"

        for cm in col_meta:
            cidx = cm["col_idx"]
            val_raw = row.iloc[cidx] if cidx < len(row) else None
            val_str = str(val_raw).strip() if val_raw is not None else ""
            if val_str.lower() in ("none", "nan", "", "n/a", "n.m.", "nm", "..", "...", "n.a", "n.a.", "-"):
                val = np.nan
            else:
                try:
                    val = float(val_raw)
                except (ValueError, TypeError):
                    val = np.nan

            records.append({
                "section": cm["section"],
                "section_key": slugify(cm["section"]),
                "country_name": country_name,
                "country_code": country_code,
                "lt_fc_rating": lt_fc_rating,
                "entity_type": entity_type,
                "indicator": cm["indicator"],
                "indicator_key": slugify(cm["indicator"]),
                "unit": cm["unit"],
                "year": cm["year"],
                "year_num": cm["year_num"],
                "is_forecast": cm["is_forecast"],
                "is_average": cm["is_average"],
                "value": val,
            })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    return df.dropna(subset=["indicator", "year"])

# ============================================================
# Excel export com gráficos (inspirado no sri_to_excel do S&P)
# ============================================================

def _fix_strref_in_zip(buf):
    """Fix numRef->strRef in cat axis and hollow markers (from S&P app)."""
    out = io.BytesIO()
    with zipfile.ZipFile(buf, 'r') as zin, \
         zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/charts/') and item.filename.endswith('.xml'):
                xml = data.decode('utf-8')
                # Fix 1: numRef -> strRef for category axis
                def _swap(m):
                    inner = (m.group(1)
                        .replace('<c:numRef>', '<c:strRef>')
                        .replace('</c:numRef>', '</c:strRef>')
                        .replace('<c:numCache>', '<c:strCache>')
                        .replace('</c:numCache>', '</c:strCache>'))
                    return '<c:cat>' + inner + '</c:cat>'
                xml = re.sub(r'<c:cat>(.*?)</c:cat>', _swap, xml, flags=re.DOTALL)
                # Fix 2: hollow markers
                NS_A = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
                NOFILL = '<a:noFill ' + NS_A + '/>'
                def _hollow_sppr(m):
                    inner = m.group(1)
                    inner = re.sub(r'<a:solidFill.*?</a:solidFill>', '', inner, flags=re.DOTALL)
                    if 'noFill' not in inner:
                        inner = NOFILL + inner
                    return '<c:spPr>' + inner + '</c:spPr>'
                def _hollow_marker(m):
                    return re.sub(r'<c:spPr>(.*?)</c:spPr>', _hollow_sppr,
                                  m.group(0), flags=re.DOTALL)
                xml = re.sub(r'<c:marker>.*?</c:marker>', _hollow_marker,
                             xml, flags=re.DOTALL)
                data = xml.encode('utf-8')
            zout.writestr(item, data)
    out.seek(0)
    return out


def _sane_sheet(name, used, ml=28):
    safe = re.sub(r'[/\\?\*\[\]:]+', '_', str(name)).strip()[:ml] or 'Sheet'
    base, n = safe, 2
    while safe in used:
        safe = base[:ml-2] + '_' + str(n)
        n += 1
    return safe


@st.cache_data(show_spinner=False)
def comparator_to_excel(df: pd.DataFrame) -> bytes:
    """Export Fitch Comparator data to Excel with line charts (inspired by S&P sri_to_excel)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    num_fmt = '#,##0.00'

    # Group by indicator
    indicators = df["indicator"].unique()
    countries = sorted(df["country_name"].unique())

    for ind in indicators:
        idf = df[df["indicator"] == ind].copy()
        if idf.empty:
            continue

        # Pivot: rows=year, columns=country
        pivot = idf.pivot_table(
            index="year", columns="country_name", values="value", aggfunc="first"
        )
        # Sort by year_num
        year_order = idf.drop_duplicates("year").set_index("year")["year_num"]
        pivot = pivot.loc[pivot.index.map(lambda y: year_order.get(y, 0)).sort_values().index]

        sheet_name = _sane_sheet(ind, used_names)
        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.cell(row=1, column=1, value=ind).font = Font(bold=True, size=12)

        # Header row (row 3)
        ws.cell(row=3, column=1, value="Year")
        ws['A3'].font = header_font_white
        ws['A3'].fill = header_fill
        ws['A3'].border = thin_border
        for ci, country in enumerate(pivot.columns, 2):
            cell = ws.cell(row=3, column=ci, value=country)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows (from row 4)
        for ri, (year_label, row_data) in enumerate(pivot.iterrows(), 4):
            ws.cell(row=ri, column=1, value=str(year_label)).border = thin_border
            for ci, country in enumerate(pivot.columns, 2):
                val = row_data.get(country)
                cell = ws.cell(row=ri, column=ci)
                if pd.notna(val):
                    cell.value = float(val)
                    cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        last_data_row = 3 + len(pivot)
        n_countries = len(pivot.columns)

        # Line chart
        if len(pivot) >= 2 and n_countries >= 1:
            chart = LineChart()
            chart.title = ind
            chart.width = 28
            chart.height = 14
            chart.style = 10
            chart.y_axis.title = ""
            chart.x_axis.title = "Year"

            # Category axis (years)
            cats = Reference(ws, min_col=1, min_row=4, max_row=last_data_row)
            chart.set_categories(cats)

            # Series
            for ci in range(2, 2 + n_countries):
                data_ref = Reference(ws, min_col=ci, min_row=3, max_row=last_data_row)
                chart.add_data(data_ref, titles_from_data=True)

            # Marker style (circle, hollow)
            for s in chart.series:
                s.graphicalProperties.line.width = 22000
                s.marker.symbol = "circle"
                s.marker.size = 5
                s.smooth = False

            chart_row = last_data_row + 2
            ws.add_chart(chart, f"A{chart_row}")

        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 25)

    # Summary sheet with all data (long format)
    summary_name = _sane_sheet("Dados_completos", used_names)
    used_names.add(summary_name)
    ws_sum = wb.create_sheet(title=summary_name, index=0)

    display_cols = ["section", "country_name", "country_code", "lt_fc_rating",
                    "indicator", "unit", "year", "value"]
    available_cols = [c for c in display_cols if c in df.columns]
    export_df = df[available_cols].sort_values(
        [c for c in ["section", "country_name", "indicator", "year_num"] if c in df.columns]
    )

    for ci, col_name in enumerate(available_cols, 1):
        cell = ws_sum.cell(row=1, column=ci, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for ri, row_data in enumerate(export_df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=ri, column=ci)
            if pd.notna(val):
                cell.value = float(val) if isinstance(val, (int, float, np.floating, np.integer)) else str(val)
            cell.border = thin_border
            if isinstance(val, (int, float, np.floating, np.integer)):
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="center")

    for col_cells in ws_sum.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws_sum.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Apply strRef fix for proper chart rendering
    fixed = _fix_strref_in_zip(buf)
    return fixed.read()

# ============================================================
# Filtros do Comparator (inspirado no build_filters do S&P)
# ============================================================

def build_comparator_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.subheader("🔍 Filtros do Comparator")

    c1, c2 = st.columns(2)

    with c1:
        entity_type = st.radio(
            "Tipo de entidade",
            ["Países", "Medianas/Grupos", "Todos"],
            horizontal=True,
            key="fc_entity_type",
        )

    if entity_type == "Países":
        df = df[df["entity_type"] == "COUNTRY"]
    elif entity_type == "Medianas/Grupos":
        df = df[df["entity_type"] != "COUNTRY"]

    with c2:
        all_ratings = sorted(df["lt_fc_rating"].dropna().unique().tolist())
        sel_ratings = st.multiselect(
            "LT FC Rating",
            options=all_ratings,
            default=[],
            key="fc_ratings",
            help="Deixe vazio para considerar todos os ratings.",
        )
    if sel_ratings:
        df = df[df["lt_fc_rating"].isin(sel_ratings)]

    c3, c4 = st.columns(2)

    with c3:
        sections = sorted(df["section"].unique())
        sel_sections = st.multiselect(
            "Seções",
            sections,
            default=sections[:3] if len(sections) >= 3 else sections,
            key="fc_sections",
        )
    if sel_sections:
        df = df[df["section"].isin(sel_sections)]

    with c4:
        countries = sorted(df["country_name"].unique())
        sel_countries = st.multiselect(
            "Países / Grupos",
            countries,
            default=[c for c in ["Brazil", "Mexico", "Colombia", "Chile", "Peru"]
                     if c in countries][:5] or countries[:5],
            key="fc_countries",
        )
    if sel_countries:
        df = df[df["country_name"].isin(sel_countries)]

    # Indicator filter
    all_indicators = sorted(df["indicator"].unique())
    sel_indicators = st.multiselect(
        "Indicadores",
        options=all_indicators,
        default=[],
        key="fc_indicators",
        help="Deixe vazio para considerar todos.",
    )
    if sel_indicators:
        df = df[df["indicator"].isin(sel_indicators)]

    # Year range slider
    valid_years = df["year_num"].dropna()
    if not valid_years.empty:
        year_min, year_max = int(valid_years.min()), int(valid_years.max())
    else:
        year_min, year_max = 2015, 2028

    c5, c6 = st.columns(2)
    with c5:
        sel_year_range = st.slider(
            "Faixa de anos",
            min_value=year_min,
            max_value=year_max,
            value=(year_min, year_max),
            key="fc_years",
        )
    df = df[df["year_num"].between(sel_year_range[0], sel_year_range[1], inclusive="both")]

    with c6:
        forecast_mode = st.radio(
            "Período",
            ["Todos", "Somente históricos", "Somente projeções"],
            index=0,
            key="fc_forecast",
            horizontal=True,
        )
    if forecast_mode == "Somente históricos":
        df = df[~df["is_forecast"]]
    elif forecast_mode == "Somente projeções":
        df = df[df["is_forecast"]]

    return df

# ============================================================
# Dashboard – gráficos de linha (inspirado no render_dashboard_tab do S&P)
# ============================================================

def render_comparator_dashboard(df: pd.DataFrame):
    st.subheader("📊 Dashboard – Fitch Comparator")

    if df.empty:
        st.warning("Nenhum dado encontrado com os filtros selecionados.")
        return

    indicators = sorted(df["indicator"].unique())

    # Exclude averages for charting
    chart_df = df[~df.get("is_average", pd.Series(False, index=df.index))].copy()
    chart_df = chart_df.sort_values("year_num")

    for ind in indicators:
        idf = chart_df[chart_df["indicator"] == ind].copy()
        if idf.empty:
            continue

        unit = idf["unit"].dropna().iloc[0] if "unit" in idf.columns and not idf["unit"].dropna().empty else ""
        title = f"{ind}" + (f" ({unit})" if unit and unit.lower() not in ("none", "nan", "") else "")

        fig = px.line(
            idf,
            x="year",
            y="value",
            color="country_name",
            markers=True,
            title=title,
            labels={"value": "", "year": "", "country_name": "País"},
        )
        fig.update_traces(mode="lines+markers")
        fig.update_layout(
            hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=-0.3),
            margin=dict(l=40, r=20, t=50, b=20),
        )
        st_plotly_chart_compat(fig)

    st.markdown("---")

    # Export buttons
    _c1, _c2 = st.columns(2)
    with _c1:
        csv_data = df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Baixar CSV",
            data=csv_data,
            file_name="fitch_comparator_filtrado.csv",
            mime="text/csv",
            key="dl_csv_dash",
        )
    with _c2:
        st.download_button(
            "⬇️ Baixar Excel (.xlsx com gráficos)",
            data=comparator_to_excel(df),
            file_name="fitch_comparator_filtrado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_dash",
        )

# ============================================================
# Dados em tabela (inspirado no render_table_tab do S&P)
# ============================================================

def render_comparator_table(df: pd.DataFrame):
    st.subheader("📋 Dados em tabela – Fitch Comparator")

    if df.empty:
        st.warning("Nenhum dado encontrado com os filtros selecionados.")
        return

    view_mode = st.radio(
        "Visualização",
        ["Longa (recomendada)", "Pivotada"],
        horizontal=True,
        index=0,
        key="fc_view_mode",
    )

    long_df = df.sort_values(
        [c for c in ["section", "country_name", "indicator", "year_num"] if c in df.columns]
    ).copy()

    if view_mode == "Longa (recomendada)":
        display_cols = [c for c in [
            "section", "country_name", "country_code", "lt_fc_rating",
            "indicator", "unit", "year", "value"
        ] if c in long_df.columns]
        display_df = long_df[display_cols]
    else:
        pivot_idx = [c for c in ["section", "country_name", "country_code",
                                  "lt_fc_rating", "indicator"] if c in long_df.columns]
        display_df = (
            long_df.pivot_table(
                index=pivot_idx,
                columns="year",
                values="value",
                aggfunc="first",
            ).reset_index()
        )

    st_dataframe_compat(display_df, use_container_width=True, hide_index=True)

    st.markdown("---")

    _c1, _c2 = st.columns(2)
    with _c1:
        csv_data = display_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Baixar CSV",
            data=csv_data,
            file_name="fitch_comparator_tabela.csv",
            mime="text/csv",
            key="dl_csv_tbl",
        )
    with _c2:
        st.download_button(
            "⬇️ Baixar Excel (.xlsx com gráficos)",
            data=comparator_to_excel(long_df),
            file_name="fitch_comparator_tabela.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_tbl",
        )

    st.caption(
        "**Legenda**: section = seção temática · indicator = indicador · "
        "year = período original · value = valor numérico · "
        "is_forecast = projeção"
    )

# ============================================================
# Metodologia – rendering (inalterado do v3)
# ============================================================

def render_methodology_overview():
    st.markdown("""
### Fitch Sovereign Rating Methodology

O modelo soberano da Fitch combina:
- **SRM (Sovereign Rating Model)**: modelo quantitativo com 18 variáveis
  agrupadas em 4 pilares, gerando um score numérico.
- **QO (Qualitative Overlay)**: ajuste qualitativo de até ±3 notches
  (extensível em crises), aplicado pilar a pilar.
- **Ratings finais**: LT FC IDR → LT LC IDR → ST FC IDR → ST LC IDR.

#### Pilares do SRM

| Pilar | Peso aprox. |
|-------|-------------|
| I. Structural Features | ~53% |
| II. Macroeconomic Performance | ~10% |
| III. Public Finances | ~19% |
| IV. External Finances | ~17% |

#### Intercepto
O intercepto OLS é **{intercept:.3f}**.

#### Escala
O score do SRM mapeia para a escala de rating usando arredondamento:
- 16 = AAA, 15 = AA+, ..., 1 = B-
- Abaixo de 1 → CCC+
    """.format(intercept=INTERCEPT))


def render_methodology_pillar(pillar_key: str):
    st.subheader(PILLAR_LABELS[pillar_key])
    st.caption("Insira os valores SRM-ready para este pilar.")

    rows = []
    for var_key, meta in SRM_VARIABLES[pillar_key].items():
        value_raw = safe_number_input(
            var_key, meta["label"],
            st.session_state.get(var_key, 0.0), meta["help"],
        )
        contribution = float(value_raw) * float(meta["coefficient"])
        rows.append({
            "Variável": meta["label"],
            "Valor": value_raw,
            "Coeficiente": meta["coefficient"],
            "Peso (%)": meta["weight"],
            "Contribuição": contribution,
        })

    rdf = pd.DataFrame(rows)
    st.dataframe(rdf, use_container_width=True, hide_index=True)
    st.metric("Subtotal do pilar", f"{rdf['Contribuição'].sum():.3f}")


def render_methodology_qo():
    st.subheader("Qualitative Overlay (QO)")
    st.caption("Ajuste qualitativo por pilar analítico. Cap típico: ±3 notches no total.")

    for pillar_key, factors in QO_FACTORS.items():
        st.markdown(f"#### {PILLAR_LABELS[pillar_key]}")
        qo_key = f"qo_{pillar_key}"

        st.selectbox(
            f"QO – {PILLAR_LABELS[pillar_key]}",
            options=list(QO_GUIDANCE.keys()),
            format_func=lambda x: f"{x:+d}  —  {QO_GUIDANCE[x]}",
            key=qo_key,
            index=2,
        )

        with st.expander("Fatores considerados"):
            for f in factors:
                st.write(f"- {f}")

    st.checkbox("Crisis extension (permite QO fora de ±3)", key="qo_crisis_extension")


def render_methodology_results():
    st.subheader("Resultados do Modelo")

    inputs = get_clean_srm_inputs()
    srm_score, details = compute_srm(inputs)

    adjustments = {
        "structural": int(st.session_state.get("qo_structural", 0)),
        "macro": int(st.session_state.get("qo_macro", 0)),
        "public_finances": int(st.session_state.get("qo_public_finances", 0)),
        "external": int(st.session_state.get("qo_external", 0)),
    }
    crisis_ext = bool(st.session_state.get("qo_crisis_extension", False))
    qo_total = clamp_qo(adjustments, crisis_ext)
    final_score = srm_score + qo_total

    lt_fc_idr = score_to_lt_rating(final_score)
    lc_adjust = int(st.session_state.get("lc_manual_adjust", 0))
    lt_lc_idr = apply_notches(lt_fc_idr, -lc_adjust)

    fc_robust = bool(st.session_state.get("fc_robust_liquidity", False))
    st_fc_idr = map_short_term(lt_fc_idr, fc_robust)
    st_lc_idr = map_short_term(lt_lc_idr, True)

    c1, c2, c3 = st.columns(3)
    c1.metric("SRM Score", f"{srm_score:.2f}")
    c2.metric("QO Total", f"{qo_total:+d}")
    c3.metric("Final Score", f"{final_score:.2f}")

    st.divider()

    r1, r2, r3, r4 = st.columns(4)
    r1.metric("LT FC IDR", lt_fc_idr)
    r2.metric("LT LC IDR", lt_lc_idr)
    r3.metric("ST FC IDR", st_fc_idr)
    r4.metric("ST LC IDR", st_lc_idr)

    st.divider()

    st.number_input(
        "LC notch adjustment vs FC (positivo = LC acima do FC)",
        min_value=-3, max_value=6, value=lc_adjust, step=1,
        key="lc_manual_adjust",
    )
    st.checkbox("FC: robust external liquidity (higher ST mapping)", key="fc_robust_liquidity")

    st.plotly_chart(build_radar(srm_score, qo_total, final_score), use_container_width=True)

    with st.expander("📋 Detalhes do SRM"):
        det_df = pd.DataFrame(details)
        st.dataframe(det_df, use_container_width=True, hide_index=True)

# ============================================================
# MAIN
# ============================================================

def main():
    init_state()

    st.sidebar.title("Fitch Sovereign App")
    st.sidebar.caption("v4 – Metodologia + Comparator (Dashboard & Tabela)")

    uploaded = st.sidebar.file_uploader(
        "📂 Carregar Fitch Comparator (.xlsb)",
        type=["xlsb"],
        key="fitch_upload",
    )

    comparator_df = None
    if uploaded is not None:
        with st.spinner("Processando arquivo XLSB…"):
            comparator_df = parse_fitch_comparator(uploaded.getvalue())
        if comparator_df is not None and not comparator_df.empty:
            st.sidebar.success(
                f"✅ {len(comparator_df):,} registros · "
                f"{comparator_df['country_name'].nunique()} entidades · "
                f"{comparator_df['indicator'].nunique()} indicadores"
            )
        else:
            st.sidebar.error("Não foi possível extrair dados do arquivo.")

    # ========== 3 TABS ==========
    tab_met, tab_dash, tab_data = st.tabs([
        "📘 Metodologia",
        "📊 Dashboard",
        "📋 Dados",
    ])

    # ========== TAB 1: Metodologia ==========
    with tab_met:
        st.title("Fitch Sovereign Rating Methodology")

        sub_page = st.radio(
            "Seção",
            [
                "Visão geral",
                PILLAR_LABELS["structural"],
                PILLAR_LABELS["macro"],
                PILLAR_LABELS["public_finances"],
                PILLAR_LABELS["external"],
                "Qualitative Overlay (QO)",
                "Resultados",
            ],
            horizontal=False,
            key="met_subpage",
        )

        if sub_page == "Visão geral":
            render_methodology_overview()
        elif sub_page == PILLAR_LABELS["structural"]:
            render_methodology_pillar("structural")
        elif sub_page == PILLAR_LABELS["macro"]:
            render_methodology_pillar("macro")
        elif sub_page == PILLAR_LABELS["public_finances"]:
            render_methodology_pillar("public_finances")
        elif sub_page == PILLAR_LABELS["external"]:
            render_methodology_pillar("external")
        elif sub_page == "Qualitative Overlay (QO)":
            render_methodology_qo()
        elif sub_page == "Resultados":
            render_methodology_results()

    # ========== TAB 2: Dashboard ==========
    with tab_dash:
        st.title("📊 Fitch Global Sovereign Data Comparator – Dashboard")
        if comparator_df is None or comparator_df.empty:
            st.info("⬅️ Envie o arquivo XLSB na barra lateral para ativar o dashboard.")
        else:
            filtered_df = build_comparator_filters(comparator_df)
            render_comparator_dashboard(filtered_df)

    # ========== TAB 3: Dados ==========
    with tab_data:
        st.title("📋 Fitch Global Sovereign Data Comparator – Dados")
        if comparator_df is None or comparator_df.empty:
            st.info("⬅️ Envie o arquivo XLSB na barra lateral para visualizar os dados.")
        else:
            # Re-apply filters for table tab (independent from dashboard)
            st.markdown("---")
            df_table = comparator_df.copy()

            tc1, tc2 = st.columns(2)
            with tc1:
                entity_type_tbl = st.radio(
                    "Tipo de entidade",
                    ["Países", "Medianas/Grupos", "Todos"],
                    horizontal=True,
                    key="fc_entity_type_tbl",
                )
            if entity_type_tbl == "Países":
                df_table = df_table[df_table["entity_type"] == "COUNTRY"]
            elif entity_type_tbl == "Medianas/Grupos":
                df_table = df_table[df_table["entity_type"] != "COUNTRY"]

            with tc2:
                all_ratings_tbl = sorted(df_table["lt_fc_rating"].dropna().unique().tolist())
                sel_ratings_tbl = st.multiselect(
                    "LT FC Rating",
                    options=all_ratings_tbl,
                    default=[],
                    key="fc_ratings_tbl",
                )
            if sel_ratings_tbl:
                df_table = df_table[df_table["lt_fc_rating"].isin(sel_ratings_tbl)]

            tc3, tc4 = st.columns(2)
            with tc3:
                sections_tbl = sorted(df_table["section"].unique())
                sel_sections_tbl = st.multiselect(
                    "Seções",
                    sections_tbl,
                    default=sections_tbl[:3] if len(sections_tbl) >= 3 else sections_tbl,
                    key="fc_sections_tbl",
                )
            if sel_sections_tbl:
                df_table = df_table[df_table["section"].isin(sel_sections_tbl)]

            with tc4:
                countries_tbl = sorted(df_table["country_name"].unique())
                sel_countries_tbl = st.multiselect(
                    "Países",
                    countries_tbl,
                    default=countries_tbl[:10] if len(countries_tbl) >= 10 else countries_tbl,
                    key="fc_countries_tbl",
                )
            if sel_countries_tbl:
                df_table = df_table[df_table["country_name"].isin(sel_countries_tbl)]

            render_comparator_table(df_table)


if __name__ == "__main__":
    main()
